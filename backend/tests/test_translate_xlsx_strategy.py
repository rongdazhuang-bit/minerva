"""Tests for XLSX translation strategy behavior."""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.translate.domain.dto import SegmentRecord
from app.translate.service.strategies.xlsx_strategy import XlsxTranslateStrategy


def test_xlsx_translates_cells_without_row_tab_join(tmp_path: Path) -> None:
    """Translate XLSX data cells individually while preserving header cells."""
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "desc"
    ws["A2"] = "apple"
    ws["B2"] = "red fruit"
    wb.save(source)

    strategy = XlsxTranslateStrategy()
    drafts = strategy.extract(source)

    assert any(
        d.anchor_json == {"sheet": "Sheet1", "row": 2, "col": 1, "label": "table_cell"}
        for d in drafts
    )
    assert any(
        d.anchor_json == {"sheet": "Sheet1", "row": 2, "col": 2, "label": "table_cell"}
        for d in drafts
    )

    records = [
        SegmentRecord(
            seq=d.seq,
            source_text=d.source_text,
            translated_text=f"译:{d.source_text}",
            anchor_json=d.anchor_json,
        )
        for d in drafts
    ]
    strategy.assemble(records, source, output)

    out = load_workbook(output)
    ws_out = out["Sheet1"]
    assert ws_out["A1"].value == "name"
    assert ws_out["B1"].value == "desc"
    assert ws_out["A2"].value == "译:apple"
    assert ws_out["B2"].value == "译:red fruit"
    out.close()
