"""File text extraction using Minerva/Dify-compatible format support."""

from __future__ import annotations

import csv
import io
from pathlib import Path

import fitz
from bs4 import BeautifulSoup
from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.exceptions import AppError


def extract_text_from_bytes(payload: bytes, *, file_name: str) -> str:
    """Extract plain text from uploaded bytes by file extension."""

    ext = Path(file_name).suffix.lower().lstrip(".")
    if ext in {"txt", "md", "markdown", "mdx", "vtt", "properties", "htm", "html"}:
        if ext in {"htm", "html"}:
            soup = BeautifulSoup(payload.decode("utf-8", errors="ignore"), "html.parser")
            return soup.get_text("\n")
        return payload.decode("utf-8", errors="ignore")
    if ext == "csv":
        text_io = io.StringIO(payload.decode("utf-8", errors="ignore"))
        reader = csv.reader(text_io)
        return "\n".join(",".join(row) for row in reader)
    if ext in {"xls", "xlsx"}:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)
    if ext == "docx":
        doc = DocxDocument(io.BytesIO(payload))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    if ext == "pdf":
        with fitz.open(stream=payload, filetype="pdf") as doc:
            return "\n".join(page.get_text() for page in doc)
    raise AppError(
        "dataset.unsupported_file_type",
        f"不支持的文件类型: .{ext}",
        422,
    )
