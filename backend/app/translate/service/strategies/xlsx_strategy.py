"""Excel xlsx document translation strategy."""

from __future__ import annotations

import uuid
from pathlib import Path
from typing import ClassVar

from openpyxl import load_workbook

from app.layout.writers.base import WriteContext
from app.layout.writers.spreadsheet_writer import XlsxCellWriter
from app.translate.domain.dto import SegmentDraft, SegmentRecord
from app.translate.service.strategies.base import DocTranslateFormatStrategy


class XlsxTranslateStrategy(DocTranslateFormatStrategy):
    """Translate ``.xlsx`` with one segment per non-empty data cell."""

    extensions: ClassVar[frozenset[str]] = frozenset({"xlsx"})

    def extract(
        self,
        local_path: Path,
        *,
        ocr_file_id: uuid.UUID | None = None,
        ocr_pages: list[tuple[int, str]] | None = None,
        layout_document=None,
    ) -> list[SegmentDraft]:
        """Extract one draft per non-empty cell, excluding the header row."""
        wb = load_workbook(local_path, read_only=True, data_only=True)
        drafts: list[SegmentDraft] = []
        seq = 0
        try:
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                for row_idx, row in enumerate(ws.iter_rows(), start=1):
                    if row_idx == 1:
                        continue
                    for col_idx, cell in enumerate(row, start=1):
                        value = "" if cell.value is None else str(cell.value)
                        if not value.strip():
                            continue
                        drafts.append(
                            SegmentDraft(
                                seq=seq,
                                source_text=value,
                                anchor_json={
                                    "sheet": sheet_name,
                                    "sheet_name": sheet_name,
                                    "page_index": sheet_idx,
                                    "row": row_idx,
                                    "col": col_idx,
                                    "label": "table_cell",
                                },
                            )
                        )
                        seq += 1
        finally:
            wb.close()
        return drafts

    def assemble(
        self,
        segments: list[SegmentRecord],
        source_path: Path,
        out_path: Path,
    ) -> None:
        """Write translated XLSX cells back to their original positions."""
        XlsxCellWriter().write(
            WriteContext(
                source_path=source_path,
                out_path=out_path,
                segments=segments,
            )
        )
