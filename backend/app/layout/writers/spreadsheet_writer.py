"""Spreadsheet writers for document translation output."""

from __future__ import annotations

from openpyxl import load_workbook

from app.layout.writers.base import WriteContext


class XlsxCellWriter:
    """Write translated text into anchored XLSX cells."""

    def write(self, context: WriteContext) -> None:
        """Apply translated segments to their anchored workbook cells."""
        wb = load_workbook(context.source_path)
        try:
            for seg in context.segments:
                anchor = seg.anchor_json or {}
                sheet = str(anchor.get("sheet", wb.sheetnames[0]))
                row = int(anchor.get("row", 1))
                col = int(anchor.get("col", 1))
                if sheet not in wb.sheetnames:
                    continue
                ws = wb[sheet]
                ws.cell(
                    row=row,
                    column=col,
                    value=seg.source_text
                    if anchor.get("skip_translate")
                    else seg.translated_text or seg.source_text,
                )
            wb.save(context.out_path)
        finally:
            wb.close()
