"""Convert local files and URLs into Markdown for PPT content pipelines."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Callable

import httpx
from bs4 import BeautifulSoup

_SUPPORTED_SUFFIXES = frozenset({".pdf", ".docx", ".xlsx", ".xlsm", ".md", ".txt"})


class IngestError(Exception):
    """Structured failure from ingest conversion with a stable machine-readable code."""

    def __init__(self, code: str, message: str) -> None:
        """Store ``code`` and human-readable ``message``."""

        super().__init__(message)
        self.code = code
        self.message = message


def convert_file_to_markdown(
    source_path: Path,
    *,
    images_dir: Path | None = None,
) -> tuple[str, list[Path]]:
    """Convert a local file to Markdown plus any extracted image paths."""

    path = Path(source_path)
    if not path.is_file():
        raise IngestError("source_missing", f"source file not found: {path}")

    suffix = path.suffix.lower()
    if suffix not in _SUPPORTED_SUFFIXES:
        raise IngestError("unsupported_format", f"unsupported file format: {suffix or '(none)'}")

    if suffix == ".pdf":
        return _convert_pdf(path, images_dir=images_dir)
    if suffix == ".docx":
        return _convert_docx(path, images_dir=images_dir)
    if suffix in {".xlsx", ".xlsm"}:
        return _convert_xlsx(path)
    return path.read_text(encoding="utf-8"), []


def convert_url_to_markdown(
    url: str,
    *,
    images_dir: Path | None = None,
) -> tuple[str, list[Path]]:
    """Fetch a URL and convert its HTML body to Markdown-like plain text."""

    del images_dir  # URL ingest does not persist remote images in Phase 2.

    target = (url or "").strip()
    if not target:
        raise IngestError("source_missing", "url is required")

    try:
        response = httpx.get(target, follow_redirects=True, timeout=30.0)
        response.raise_for_status()
    except httpx.HTTPError as exc:
        raise IngestError("convert_failed", f"failed to fetch url: {exc}") from exc

    content_type = response.headers.get("content-type", "").lower()
    body = response.text
    if "html" in content_type or _looks_like_html(body):
        markdown = _html_to_markdown(body)
    else:
        markdown = body.strip()

    if not markdown.strip():
        raise IngestError("convert_failed", "url returned empty content")

    return markdown, []


def build_image_manifest(image_paths: list[Path], base_dir: Path) -> dict:
    """Build a JSON-serializable manifest for extracted images under ``base_dir``."""

    entries: list[dict[str, object]] = []
    for image_path in image_paths:
        try:
            rel = image_path.resolve().relative_to(base_dir.resolve())
            rel_str = rel.as_posix()
        except ValueError:
            rel_str = image_path.name
        size_bytes = image_path.stat().st_size if image_path.is_file() else 0
        entries.append(
            {
                "path": rel_str,
                "filename": image_path.name,
                "size_bytes": size_bytes,
            }
        )
    return {"images": entries, "count": len(entries)}


def _convert_pdf(source_path: Path, *, images_dir: Path | None) -> tuple[str, list[Path]]:
    """Extract per-page text and optional embedded images from a PDF."""

    try:
        import fitz
    except ImportError as exc:
        raise IngestError("convert_failed", "PyMuPDF (fitz) is not installed") from exc

    image_paths: list[Path] = []
    parts: list[str] = []
    if images_dir is not None:
        images_dir.mkdir(parents=True, exist_ok=True)

    try:
        doc = fitz.open(source_path)
    except Exception as exc:
        raise IngestError("convert_failed", f"failed to open pdf: {exc}") from exc

    try:
        for page_index in range(len(doc)):
            page = doc[page_index]
            text = page.get_text().strip()
            parts.append(f"## Page {page_index + 1}\n\n{text}" if text else f"## Page {page_index + 1}")

            if images_dir is None:
                continue
            for img_index, image_info in enumerate(page.get_images(full=True)):
                xref = image_info[0]
                try:
                    extracted = doc.extract_image(xref)
                except Exception:
                    continue
                ext = extracted.get("ext") or "png"
                image_bytes = extracted.get("image")
                if not image_bytes:
                    continue
                image_path = images_dir / f"page{page_index + 1}_img{img_index + 1}.{ext}"
                image_path.write_bytes(image_bytes)
                image_paths.append(image_path)
    finally:
        doc.close()

    markdown = "\n\n".join(parts).strip()
    if not markdown:
        raise IngestError("convert_failed", "pdf contains no extractable text")
    return markdown, image_paths


def _convert_docx(source_path: Path, *, images_dir: Path | None) -> tuple[str, list[Path]]:
    """Convert DOCX to Markdown via mammoth, optionally saving embedded images."""

    try:
        import mammoth
    except ImportError as exc:
        raise IngestError("convert_failed", "mammoth is not installed") from exc

    image_paths: list[Path] = []
    if images_dir is not None:
        images_dir.mkdir(parents=True, exist_ok=True)

    convert_image: Callable | None = None
    if images_dir is not None:
        counter = {"n": 0}

        def _save_image(image: object) -> dict[str, str]:
            counter["n"] += 1
            ext = "png"
            content_type = getattr(image, "content_type", "") or ""
            if "jpeg" in content_type or "jpg" in content_type:
                ext = "jpg"
            elif "gif" in content_type:
                ext = "gif"
            image_path = images_dir / f"image_{counter['n']}.{ext}"
            with image.open() as image_bytes:  # type: ignore[union-attr]
                image_path.write_bytes(image_bytes.read())
            image_paths.append(image_path)
            return {"src": image_path.name}

        convert_image = mammoth.images.img_element(_save_image)

    try:
        with source_path.open("rb") as docx_file:
            if convert_image is not None:
                result = mammoth.convert_to_markdown(docx_file, convert_image=convert_image)
            else:
                result = mammoth.convert_to_markdown(docx_file)
    except Exception as exc:
        raise IngestError("convert_failed", f"failed to convert docx: {exc}") from exc

    markdown = (result.value or "").strip()
    if not markdown:
        raise IngestError("convert_failed", "docx contains no extractable text")
    return markdown, image_paths


def _convert_xlsx(source_path: Path) -> tuple[str, list[Path]]:
    """Render each worksheet as a Markdown table."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngestError("convert_failed", "openpyxl is not installed") from exc

    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError("convert_failed", f"failed to open spreadsheet: {exc}") from exc

    parts: list[str] = []
    try:
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows = [
                ["" if cell is None else str(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            if not rows:
                continue
            parts.append(f"## {sheet_name}\n")
            header = rows[0]
            parts.append("| " + " | ".join(header) + " |")
            parts.append("| " + " | ".join("---" for _ in header) + " |")
            for row in rows[1:]:
                padded = row + [""] * (len(header) - len(row))
                parts.append("| " + " | ".join(padded[: len(header)]) + " |")
    finally:
        workbook.close()

    markdown = "\n\n".join(parts).strip()
    if not markdown:
        raise IngestError("convert_failed", "spreadsheet contains no data")
    return markdown, []


def _looks_like_html(text: str) -> bool:
    """Return True when ``text`` appears to contain HTML markup."""

    snippet = text.lstrip()[:500].lower()
    return snippet.startswith("<!doctype html") or snippet.startswith("<html") or "<body" in snippet


def _html_to_markdown(html: str) -> str:
    """Convert HTML to readable plain text with basic heading structure."""

    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    title = soup.find("title")
    parts: list[str] = []
    if title and title.get_text(strip=True):
        parts.append(f"# {title.get_text(strip=True)}")

    body = soup.body or soup
    for element in body.find_all(["h1", "h2", "h3", "h4", "h5", "h6", "p", "li"]):
        text = element.get_text(" ", strip=True)
        if not text:
            continue
        name = element.name or "p"
        if name.startswith("h") and len(name) == 2 and name[1].isdigit():
            level = int(name[1])
            parts.append(f"{'#' * level} {text}")
        elif name == "li":
            parts.append(f"- {text}")
        else:
            parts.append(text)

    if not parts:
        text = body.get_text("\n", strip=True)
        parts = [line for line in text.splitlines() if line.strip()]

    cleaned = "\n\n".join(parts)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()
